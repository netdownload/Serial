import pymysql
import openpyxl
import datetime
from pathlib import Path


DATABASE_HOST = '10.1.1.99'
DATABASE_USER = 'user'
DATABASE_PASSWORD = 'qwerty123'
DATABASE = 'resources'

date_time_begin = '2021-12-17 00:00:00'
date_time_begin_obj = datetime.datetime.strptime(date_time_begin, '%Y-%m-%d %H:%M:%S')
# Функция записывает считаные данные в базу
def insert_values_into_database55(electro_datetime, active_power, reactive_power):
    connection = pymysql.connect(host=DATABASE_HOST,
                                 user=DATABASE_USER,
                                 password=DATABASE_PASSWORD,
                                 db=DATABASE)
    try:
        with connection.cursor() as cursor:
            sql = "INSERT INTO `electro55` (`electro55_datetime`, `electro55_active`, `electro55_reactive`) VALUES (" \
                  "%s, %s, %s)"
            cursor.execute(sql, (electro_datetime, active_power, reactive_power))
            connection.commit()
    finally:
        connection.close()

def insert_values_into_database56(electro_datetime, active_power, reactive_power):
    connection = pymysql.connect(host=DATABASE_HOST,
                                 user=DATABASE_USER,
                                 password=DATABASE_PASSWORD,
                                 db=DATABASE)
    try:
        with connection.cursor() as cursor:
            sql = "INSERT INTO `electro56` (`electro56_datetime`, `electro56_active`, `electro56_reactive`) VALUES (" \
                  "%s, %s, %s)"
            cursor.execute(sql, (electro_datetime, active_power, reactive_power))
            connection.commit()
    finally:
        connection.close()


xlsx_file = Path('to_db', '333.xlsx')
wb_obj = openpyxl.load_workbook(xlsx_file)

# Read the active sheet:
sheet = wb_obj.active
for row in sheet.iter_rows(10, sheet.max_row-1):
    # electro_datetime = row[0].value[6:10] + "-" + row[0].value[3:5] + "-" + row[0].value[0:2] + row[0].value[10:16]
    date_time_begin_obj = date_time_begin_obj + datetime.timedelta(minutes=30)
    print(date_time_begin_obj)
    print(round((row[1].value/120), 4))
    active_power55 = round((row[1].value/120), 4)
    print(round((row[3].value / 120), 4))
    reactive_power55 = round((row[3].value/120), 4)

    print(round((row[5].value/120), 4))
    active_power56 = round((row[5].value/120), 4)
    print(round((row[7].value / 120), 4))
    reactive_power56 = round((row[7].value/120), 4)

    insert_values_into_database55(date_time_begin_obj, active_power55, reactive_power55)
    insert_values_into_database56(date_time_begin_obj, active_power56, reactive_power56)

    date_time_begin_obj = date_time_begin_obj + datetime.timedelta(minutes=30)
    print(date_time_begin_obj)
    print(round((row[1].value/120), 4))
    active_power55 = round((row[1].value/120), 4)
    print(round((row[3].value / 120), 4))
    reactive_power55 = round((row[3].value/120), 4)

    print(round((row[5].value/120), 4))
    active_power56 = round((row[5].value/120), 4)
    print(round((row[7].value / 120), 4))
    reactive_power56 = round((row[7].value/120), 4)

    insert_values_into_database55(date_time_begin_obj, active_power55, reactive_power55)
    insert_values_into_database56(date_time_begin_obj, active_power56, reactive_power56)

